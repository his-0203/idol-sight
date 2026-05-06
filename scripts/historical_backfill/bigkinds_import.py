#!/usr/bin/env python3
# /// script
# requires-python = ">=3.12"
# dependencies = ["openpyxl>=3.1"]
# ///
"""BIGKinds CSV/Excel import → weekly news count → SQL UPSERT.

Naver Open API 의 1000-article cap + 날짜 범위 필터 부재로 회수 못한
historical 뉴스 카운트를 사용자가 BIGKinds 웹사이트에서 직접 검색 +
CSV/Excel export 한 후 이 스크립트가 weekly bucket → migration 생성.

## 사용자 작업 (8 그룹 × 한 번씩)

1. https://www.bigkinds.or.kr/v2/news/index.do 접속
2. 검색창에 그룹 키워드 입력 (아래 표 참조)
3. 좌측 필터 → "기간" → 직접 입력 (그룹별 backfill 창 — 아래 표)
4. 좌측 필터 → 매체 (옵션, 선택 사항 — 모든 매체 권장)
5. 검색 실행
6. 결과 화면 우측 상단 **"분석" → "STEP 03 데이터 다운로드" → "엑셀 다운로드"** 또는 검색 결과 페이지의 다운로드 버튼
7. 받은 .xlsx 또는 .csv 파일을 다음 경로에 저장:
     scripts/historical_backfill/bigkinds/<group_key>.xlsx
   (또는 .csv)
8. 8 그룹 모두 끝나면 이 스크립트 실행

## 그룹별 검색 조건

| group_key | 검색 키워드 | 기간 (D-180 ~ 현재) |
|-----------|-------------|---------------------|
| plave     | "플레이브" OR "PLAVE"     | 2022-09-13 ~ 현재  |
| skinz     | "SKINZ" OR "스킨즈"       | 2024-10-12 ~ 현재  |
| myrakl    | "MY:RAKL" OR "마이라클"   | 2025-07-30 ~ 현재  |
| owis     | "OWIS" OR "오위스"        | 2025-09-24 ~ 현재  |
| miiwan   | "MiiWAN" OR "미이완"      | 2025-12-01 ~ 현재  |
| bdawn    | "B:DAWN" OR "비던"        | 2025-11-04 ~ 현재  |
| isedol   | "이세계아이돌" OR "ISEDOL" | 2021-06-17 ~ 현재  |
| stellive | "스텔라이브" OR "STELLIVE" | 2022-04-16 ~ 현재  |

⚠ BIGKinds 무료 다운로드는 1회 최대 5000건. 5000건 초과 시 기간을 분할
해 여러 번 다운로드, 같은 디렉토리에 _1.xlsx, _2.xlsx 형식으로 저장 →
이 스크립트가 모든 .xlsx/.csv 자동 인식.

## 실행

  uv run scripts/historical_backfill/bigkinds_import.py

→ migrations/0030_bigkinds_news_backfill.sql 생성.
"""

from __future__ import annotations

import csv
import sys
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterator

ROOT = Path(__file__).resolve().parent.parent.parent
INPUT_DIR = ROOT / "scripts" / "historical_backfill" / "bigkinds"
MIGRATION_PATH = ROOT / "migrations" / "0030_bigkinds_news_backfill.sql"

# 8 그룹 — naver_api_fetch.py 와 동기. group_key 는 디렉토리 내 파일명
# 접두사로 매칭 (e.g., plave.xlsx, plave_1.xlsx 모두 plave 로 인식).
GROUPS = ["plave", "skinz", "myrakl", "owis",
          "miiwan", "bdawn", "isedol", "stellive"]


def _parse_date(raw: str) -> date | None:
    """BIGKinds 의 date 컬럼은 다양한 포맷:
    'YYYYMMDD', 'YYYY-MM-DD', 'YYYY.MM.DD', 'YYYY/MM/DD' — 모두 처리.
    """
    if not raw:
        return None
    s = raw.strip()
    # 숫자 8자리 = YYYYMMDD
    if len(s) == 8 and s.isdigit():
        try:
            return date(int(s[0:4]), int(s[4:6]), int(s[6:8]))
        except ValueError:
            return None
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _read_xlsx(path: Path) -> Iterator[dict[str, Any]]:
    from openpyxl import load_workbook  # type: ignore[import-untyped]
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return
    cols = [str(c) if c is not None else "" for c in header]
    for r in rows:
        yield dict(zip(cols, r))


def _read_csv(path: Path) -> Iterator[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            yield row


def _detect_date_column(headers: list[str]) -> str | None:
    """BIGKinds 컬럼명은 한국어 기반: '일자', '날짜', '기사일자', 'date' 등."""
    candidates = ["일자", "날짜", "기사일자", "발행일", "date", "Date"]
    for cand in candidates:
        for h in headers:
            if cand in h:
                return h
    return None


def bucket_group(group_key: str) -> dict[date, int]:
    """디렉토리 내 group_key 접두사 파일들을 모두 읽어 weekly bucket."""
    counter: Counter[date] = Counter()
    if not INPUT_DIR.exists():
        return {}
    files = sorted(INPUT_DIR.glob(f"{group_key}*.xlsx")) + \
            sorted(INPUT_DIR.glob(f"{group_key}*.csv"))
    if not files:
        return {}
    for f in files:
        rows = _read_xlsx(f) if f.suffix == ".xlsx" else _read_csv(f)
        first = next(iter(rows), None)
        if first is None:
            continue
        date_col = _detect_date_column(list(first.keys()))
        if date_col is None:
            print(f"  ⚠ {f.name}: 날짜 컬럼 자동 감지 실패. "
                  f"컬럼: {list(first.keys())[:5]}", file=sys.stderr)
            continue
        # 첫 행 처리 + 나머지
        row_iter = (
            r for r in [first, *(_read_xlsx(f) if f.suffix == ".xlsx" else _read_csv(f))]
        )
        # 위 generator 가 두 번 enumerate 하므로 다시 read
        rows2 = _read_xlsx(f) if f.suffix == ".xlsx" else _read_csv(f)
        for r in rows2:
            d = _parse_date(str(r.get(date_col, "")))
            if d is None:
                continue
            counter[_week_start(d)] += 1
    return dict(counter)


def emit_sql_block(group_key: str, weekly: dict[date, int]) -> str:
    if not weekly:
        return f"-- {group_key.upper()}: 입력 파일 없음 또는 비어있음 — skip\n"
    rows = sorted(weekly.items())
    values = ",\n  ".join(
        f"('{group_key}', '{d.isoformat()}T00:00:00Z', NULL, NULL, NULL, "
        f"0, 0, 0, {n}, 0, 0, 'backfill_exact')"
        for d, n in rows
    )
    return f"""-- ============================================================
-- {group_key.upper()} BIGKinds weekly ({len(rows)} weeks, {rows[0][0].isoformat()} ~ {rows[-1][0].isoformat()})
-- Source: BIGKinds 검색 결과 사용자 다운로드 → bucketed pubDate
-- ============================================================
INSERT INTO agg_summary
  (group_key, snapshot_at,
   yt_total_videos, yt_total_views, yt_subscribers,
   dc_total_posts, theqoo_posts, instiz_posts,
   naver_total_news, twitter_posts, controversy_count, data_source)
VALUES
  {values}
ON CONFLICT(group_key, snapshot_at) DO UPDATE SET
  naver_total_news = excluded.naver_total_news,
  data_source      = CASE
    WHEN agg_summary.data_source = 'live' THEN 'live'
    ELSE excluded.data_source
  END;
"""


def main() -> None:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"BIGKinds CSV/XLSX 입력 디렉토리: {INPUT_DIR}", file=sys.stderr)
    print(f"파일 패턴: <group_key>*.xlsx 또는 <group_key>*.csv\n", file=sys.stderr)

    blocks: list[str] = []
    summary: list[tuple[str, int, int]] = []
    for gk in GROUPS:
        weekly = bucket_group(gk)
        blocks.append(emit_sql_block(gk, weekly))
        summary.append((gk, len(weekly), sum(weekly.values())))

    MIGRATION_PATH.parent.mkdir(parents=True, exist_ok=True)
    with MIGRATION_PATH.open("w", encoding="utf-8") as f:
        f.write("-- migrations/0030_bigkinds_news_backfill.sql\n")
        f.write("-- BIGKinds 검색 결과 사용자 다운로드 → weekly news count.\n")
        f.write("-- data_source='backfill_exact' (authoritative, BIGKinds = 한국 언론사 통합 archive).\n")
        f.write("-- Generated by scripts/historical_backfill/bigkinds_import.py\n\n")
        for blk in blocks:
            f.write(blk + "\n")

    print("=== summary ===", file=sys.stderr)
    for gk, weeks, total in summary:
        status = "OK" if weeks > 0 else "(no input file)"
        print(f"  {gk:<10} weeks={weeks:>4}  total_articles={total:>6}  {status}",
              file=sys.stderr)
    print(f"\nMigration 작성됨: {MIGRATION_PATH}", file=sys.stderr)


if __name__ == "__main__":
    main()
